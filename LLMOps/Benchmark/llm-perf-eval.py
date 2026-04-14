#!/usr/bin/env python3
"""LLM 性能评测工具

用法:
    python llm-perf-eval.py -b bench-conf.yaml -m model-conf.yaml

详见 README.md
"""
import ast
import argparse
import logging
import os
import re
import shlex
import subprocess
import sys
import time
from datetime import datetime

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from evalscope.perf.main import run_perf_benchmark
from evalscope.perf.arguments import Arguments


_EXCEL_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _sanitize_for_excel(text):
    """移除 Excel/openpyxl 不支持的非法控制字符"""
    return _EXCEL_ILLEGAL_CHARS_RE.sub("", text)


# ============================================================
# 日志
# ============================================================

def get_logger(level=logging.INFO):
    _logger = logging.getLogger("LLM-PerfEval")
    _logger.setLevel(level)
    _logger.propagate = False
    if not _logger.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setLevel(level)
        handler.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
        ))
        _logger.addHandler(handler)
    return _logger


logger = get_logger(logging.INFO)


# ============================================================
# 工具函数
# ============================================================

def parse_time(val):
    """解析时间值，支持 s、ms 单位或纯数字（默认秒）"""
    if isinstance(val, str):
        val = val.strip()
        if val.endswith("ms"):
            return float(val[:-2]) / 1000.0
        elif val.endswith("s"):
            return float(val[:-1])
        else:
            return float(val)
    return float(val)


def parse_model_cmd_args(cmd_str):
    """
    解析模型启动命令中的参数。

    规则：
    - 支持参数出现在命令任意位置（不依赖顺序）
    - 参数以 - 或 -- 开头，去除前缀
    - 值可以用 = 或空格分隔
    - 无值参数视为布尔标志 (true)
    - 引号包裹的大字段（如 JSON 字符串）由 shlex 负责拆分，整体作为一个值

    返回: dict，key=参数名, value=参数值(str) 或 "true"
    """
    cmd_str = cmd_str.replace("\\\n", " ")
    tokens = shlex.split(cmd_str)

    params = {}

    def _is_option_token(tok):
        if not tok or tok == "-" or tok == "--":
            return False
        if not tok.startswith("-"):
            return False
        # 仅将 -x / -abc / --key / --key-name / --key.name 视为参数 token
        # 避免把负数（如 -1）误判为参数
        return re.match(r"^-{1,2}[A-Za-z_][A-Za-z0-9_.-]*$", tok) is not None

    def _is_value_token(tok):
        # 只要不是新的参数 token，就当作上一个参数的值
        return not _is_option_token(tok)

    i = 0
    while i < len(tokens):
        token = tokens[i]

        if not _is_option_token(token):
            i += 1
            continue

        # --key=value 或 -key=value
        if "=" in token:
            key, value = token.split("=", 1)
            key = key.lstrip("-")
            params[key] = value
            i += 1
            continue

        key = token.lstrip("-")

        # 下一个 token 不是参数则作为值，否则布尔标志
        if i + 1 < len(tokens) and _is_value_token(tokens[i + 1]):
            params[key] = tokens[i + 1]
            i += 2
        else:
            params[key] = "true"
            i += 1

    return params


def apply_override_args(model_cmd, override_args):
    """在模型启动命令上应用参数覆盖。

    override_args: 参数覆盖列表
      - "--key value"   → 添加或覆盖带值参数
      - "--flag"        → 添加或覆盖布尔标志
      - "!--key"        → 删除参数（带值或布尔标志均可）

    返回: 应用覆盖后的命令字符串
    """
    if not override_args:
        return model_cmd

    def _is_option(tok):
        if not tok or tok in ('-', '--'):
            return False
        if not tok.startswith('-'):
            return False
        return re.match(r'^-{1,2}[A-Za-z_][A-Za-z0-9_.-]*$', tok) is not None

    normalized = model_cmd.replace("\\\n", " ")
    tokens = shlex.split(normalized)

    prefix = []
    args = []           # [[key_with_dashes, value_or_None], ...]
    args_key_map = {}   # key_name -> index in args

    i = 0
    while i < len(tokens):
        tok = tokens[i]
        if _is_option(tok):
            key_name = tok.lstrip('-')
            if i + 1 < len(tokens) and not _is_option(tokens[i + 1]):
                args.append([tok, tokens[i + 1]])
                args_key_map[key_name] = len(args) - 1
                i += 2
            else:
                args.append([tok, None])
                args_key_map[key_name] = len(args) - 1
                i += 1
        else:
            prefix.append(tok)
            i += 1

    for override in override_args:
        override = override.strip()
        if not override:
            continue

        if override.startswith('!'):
            # 删除参数
            key_part = override[1:].strip()
            key_name = key_part.lstrip('-')
            if key_name in args_key_map:
                idx = args_key_map.pop(key_name)
                args[idx] = None
            else:
                logger.warning(f"override_args 删除目标不存在, 已忽略: {override}")
        else:
            # 添加或覆盖
            override_tokens = shlex.split(override)
            if not override_tokens:
                continue
            key_tok = override_tokens[0]
            key_name = key_tok.lstrip('-')
            value = override_tokens[1] if len(override_tokens) > 1 else None

            if key_name in args_key_map:
                idx = args_key_map[key_name]
                args[idx] = [key_tok, value]
            else:
                args.append([key_tok, value])
                args_key_map[key_name] = len(args) - 1

    parts = prefix[:]
    for arg in args:
        if arg is None:
            continue
        parts.append(arg[0])
        if arg[1] is not None:
            parts.append(arg[1])

    return ' '.join(parts)


def _cmds_are_equivalent(cmd1, cmd2):
    """判断两个模型启动命令的参数是否等价（忽略参数顺序）"""
    return parse_model_cmd_args(cmd1) == parse_model_cmd_args(cmd2)


def collect_all_param_keys(commands_params, exclude=None):
    """收集所有命令中出现的参数名，保持出现顺序，过滤排除项"""
    exclude = set(exclude or [])
    seen = set()
    keys = []
    for params in commands_params:
        for k in params:
            if k not in seen and k not in exclude:
                seen.add(k)
                keys.append(k)
    return keys


# ============================================================
# 配置加载
# ============================================================

def load_yaml(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# bench.yaml 默认值
BENCH_DEFAULTS = {
    "mode": 1,
    "bench_only": False,
    "bench_only_healthcheck": False,
    "api_key": "",
    "dataset": "random",
    "result_dir": "./results",
    "excel_prefix": "",
    "warmup": {"enable": False},
    "healthcheck": {
        "initial_delay": 120,
        "interval": 5,
        "retry_count": 60,
    },
}

# model.yaml 默认值
MODEL_DEFAULTS = {
    "stop_cmd": "docker stop model-server && docker rm -f model-server",
}

# SLO 默认值
SLO_DEFAULTS = {
    "search_method": "binary",
    "stop_n": 2,
    "init_concurrent": 2,
    "request_multiplier": 4,
    "max_concurrent": None,
}

# healthcheck 默认值
HEALTHCHECK_DEFAULTS = {
    "initial_delay": 120,
    "interval": 5,
    "retry_count": 60,
}


def _apply_defaults(conf, defaults):
    """将 defaults 中的键值填充到 conf 中（仅填充 conf 中不存在的键）"""
    for k, v in defaults.items():
        if k not in conf or conf[k] is None:
            conf[k] = v
    return conf


def load_bench_config(file_path):
    """加载压测配置文件"""
    conf = load_yaml(file_path)
    _apply_defaults(conf, BENCH_DEFAULTS)
    mode = conf["mode"]

    for key in ("model_name", "tokenizer_path", "url"):
        if not conf.get(key):
            raise ValueError(f"bench 配置缺少必填字段: {key}")

    # healthcheck 子字段默认值
    if isinstance(conf.get("healthcheck"), dict):
        _apply_defaults(conf["healthcheck"], HEALTHCHECK_DEFAULTS)

    if mode == 1:
        conf["test_cases"] = _parse_mode1_test_cases(conf.get("test_cases", {}))
    elif mode == 2:
        conf["slo"] = _parse_mode2_slo(conf.get("slo", {}))
    else:
        raise ValueError(f"不支持的测试模式: {mode}")

    # result_file_prefix 默认由 model_name 生成，忽略开头的 /，中间的 / 替换为 --
    if not conf.get("result_file_prefix"):
        name = conf["model_name"].strip("/")
        conf["result_file_prefix"] = name.replace("/", "--")

    # 可选 Excel 自定义前缀：非空时追加在现有前缀前面
    custom_excel_prefix = str(conf.get("excel_prefix", "") or "").strip()
    if custom_excel_prefix:
        conf["result_file_prefix"] = f"{custom_excel_prefix}-{conf['result_file_prefix']}"

    logger.info(f"bench 配置加载完成 (mode={mode}, warmup={conf['warmup']})")
    return conf


def _parse_mode1_test_cases(tc):
    """解析模式1的测试用例配置"""
    context = [ast.literal_eval(c) if isinstance(c, str) else c for c in tc.get("context", [])]
    batch_request = [ast.literal_eval(b) if isinstance(b, str) else b for b in tc.get("batch_request", [])]
    combination_mode = tc.get("combination_mode", 1)

    if combination_mode == 1:
        cases = [(ctx, br) for ctx in context for br in batch_request]
    elif combination_mode == 2:
        if len(context) != len(batch_request):
            raise ValueError("combination_mode=2 时 context 与 batch_request 长度必须一致")
        cases = list(zip(context, batch_request))
    else:
        raise ValueError(f"不支持的 combination_mode: {combination_mode}")

    return {"cases": cases, "combination_mode": combination_mode}


def _parse_mode2_slo(slo):
    """解析模式2的SLO配置"""
    _apply_defaults(slo, SLO_DEFAULTS)
    raw_criteria = slo.get("criteria", {})
    supported = {"ttft", "tpot"}
    parsed_criteria = {}
    for metric, threshold in raw_criteria.items():
        if metric not in supported:
            logger.warning(f"忽略不支持的 criteria 指标: {metric}")
            continue
        parsed_criteria[metric] = parse_time(threshold)

    if not parsed_criteria:
        raise ValueError("SLO criteria 未配置任何有效指标(支持: ttft, tpot)")

    criteria_desc = ", ".join([f"{k}<={v}s" for k, v in parsed_criteria.items()])
    logger.info(f"生效指标准则: {criteria_desc}")

    raw_contexts = slo.get("context", [])
    contexts = []
    for item in raw_contexts:
        if isinstance(item, dict):
            ctx_val = item.get("context")
            if isinstance(ctx_val, str):
                ctx_val = ast.literal_eval(ctx_val)
            override_args = item.get("override_args") or []
            engine_override_args = item.get("engine_override_args") or {}
            contexts.append({
                "context": ctx_val,
                "init_concurrent": item.get("init_concurrent"),
                "max_concurrent": item.get("max_concurrent"),
                "override_args": override_args,
                "engine_override_args": engine_override_args,
            })
        else:
            ctx_val = ast.literal_eval(item) if isinstance(item, str) else item
            contexts.append({"context": ctx_val, "init_concurrent": None, "max_concurrent": None, "override_args": [], "engine_override_args": {}})

    slo["criteria_parsed"] = parsed_criteria
    slo["contexts"] = contexts
    return slo


def load_model_config(file_path):
    """加载模型启动配置文件，支持多推理引擎"""
    conf = load_yaml(file_path)
    top_level_ssh = conf.get("ssh_cmd")
    top_level_exclude = conf.get("exclude_params", [])

    raw_engines = conf.get("engines")
    if not isinstance(raw_engines, list) or not raw_engines:
        raise ValueError("model 配置缺少 engines 列表或为空")

    engines = []
    for idx, eng in enumerate(raw_engines, 1):
        if not isinstance(eng, dict):
            raise ValueError(f"engines[{idx}] 配置格式错误，必须是 dict")

        _apply_defaults(eng, MODEL_DEFAULTS)

        engine_name = eng.get("engine", f"engine-{idx}")
        if not eng.get("container_cmd"):
            raise ValueError(f"engines[{idx}]({engine_name}) 缺少 container_cmd")
        if not eng.get("commands"):
            raise ValueError(f"engines[{idx}]({engine_name}) 缺少 commands")

        eng["engine_name"] = engine_name
        eng["container_cmd"] = eng["container_cmd"].strip()
        eng["stop_cmd"] = eng["stop_cmd"].strip()
        for item in eng["commands"]:
            item["cmd"] = item["cmd"].strip()

        # 从 container_cmd 中解析容器名称
        eng["container_name"] = _parse_container_name(eng["container_cmd"])

        # 提取镜像名称
        eng["image_name"] = _parse_image_name(eng["container_cmd"])

        # exclude_params：合并顶层 + engine 级（取并集）
        engine_exclude = eng.get("exclude_params") or []
        eng["exclude_params"] = list(dict.fromkeys(top_level_exclude + engine_exclude))

        # ssh_cmd：engine 级优先，否则继承顶层
        ssh_cmd = eng.get("ssh_cmd", top_level_ssh)
        eng["ssh_cmd"] = ssh_cmd.strip() if ssh_cmd else None

        logger.info(
            f"加载推理引擎[{idx}]: {engine_name} | 镜像: {eng['image_name']} | "
            f"命令组数: {len(eng['commands'])}"
        )

        engines.append(eng)

    conf["engines"] = engines
    return conf


def _parse_container_name(container_cmd):
    """从 docker run 命令中解析 --name 参数"""
    match = re.search(r'--name\s+(\S+)', container_cmd)
    return match.group(1) if match else None


def _parse_image_name(container_cmd):
    """从 docker run 命令中提取镜像名称（最后一个非选项 token）"""
    cmd = container_cmd.replace("\\\n", " ")
    tokens = shlex.split(cmd)
    return tokens[-1] if tokens else container_cmd


# ============================================================
# 模型容器生命周期管理
# ============================================================

class ModelManager:
    def __init__(self, model_config, bench_config):
        self.container_cmd = model_config["container_cmd"]
        self.stop_cmd = model_config["stop_cmd"]
        self.ssh_cmd = model_config.get("ssh_cmd")
        self.container_name = model_config.get("container_name")
        self.url = bench_config["url"]
        self.model_name = bench_config["model_name"]
        self.healthcheck = bench_config["healthcheck"]

    def _run_cmd(self, cmd, timeout=120):
        """执行命令，如果配置了 ssh_cmd 则通过 SSH 远程执行"""
        if self.ssh_cmd:
            # 用单引号包裹远程命令，内部单引号转义
            escaped = cmd.replace("'", "'\"'\"'")
            exec_cmd = f"{self.ssh_cmd} '{escaped}'"
        else:
            exec_cmd = cmd
        return subprocess.run(
            exec_cmd, shell=True, capture_output=True, text=True, timeout=timeout
        )

    def build_full_cmd(self, model_cmd):
        """拼接容器启动命令 + 模型启动命令"""
        entrypoint_shell = re.search(r"--entrypoint\s+(?:/bin/)?(?:bash|sh)\b", self.container_cmd)
        if entrypoint_shell:
            # 当 entrypoint 是 shell 时，必须通过 -lc 执行模型命令；
            # 否则会被 bash/sh 当作脚本文件解释，导致 python 命令报 import/from 语法错误。
            return f"{self.container_cmd} -lc {shlex.quote(model_cmd)}"
        return f"{self.container_cmd} {model_cmd}"

    def cleanup(self):
        """清理残留容器（忽略错误，容器可能不存在）"""
        logger.info("清理残留容器...")
        try:
            result = self._run_cmd(self.stop_cmd, timeout=60)
            if result.returncode == 0:
                logger.info("残留容器已清理")
            else:
                logger.debug(f"无残留容器或已清理: {result.stderr.strip()}")
        except Exception:
            pass
        # 等待一下确保资源释放
        time.sleep(2)

    def start_model(self, model_cmd):
        """启动模型服务容器，返回 (success, logs)"""
        # 先清理残留容器，避免上次中断后重启报错
        self.cleanup()

        full_cmd = self.build_full_cmd(model_cmd)
        location = f"远程({self.ssh_cmd.split()[-1]})" if self.ssh_cmd else "本地"
        logger.info(f"启动模型服务 [{location}]...\n命令: {full_cmd}")
        try:
            result = self._run_cmd(full_cmd, timeout=120)
            if result.returncode == 0:
                logger.info("容器启动命令执行成功")
                return True, result.stdout
            else:
                container_logs = self.get_container_logs()
                logs = f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}\n容器日志:\n{container_logs}"
                logger.error(f"容器启动失败 (returncode={result.returncode})\n{logs}")
                return False, logs
        except subprocess.TimeoutExpired:
            logger.error("容器启动命令执行超时")
            return False, "启动命令执行超时(120s)"
        except Exception as e:
            logger.error(f"容器启动异常: {e}")
            return False, str(e)

    def stop_model(self):
        """停止模型服务容器"""
        logger.info(f"停止模型服务...\n命令: {self.stop_cmd}")
        try:
            result = self._run_cmd(self.stop_cmd, timeout=60)
            if result.returncode == 0:
                logger.info("模型服务已停止")
            else:
                logger.warning(f"停止命令返回非零: {result.stderr.strip()}")
        except Exception as e:
            logger.warning(f"停止模型服务异常: {e}")

    def health_check(self):
        """对模型服务进行健康检查（始终从本机发起 HTTP 请求）"""
        initial_delay = self.healthcheck.get("initial_delay", 120)
        interval = self.healthcheck.get("interval", 5)
        retry_count = self.healthcheck.get("retry_count", 60)

        logger.info(f"等待 {initial_delay}s 后开始健康检查...")
        time.sleep(initial_delay)

        headers = {"Content-Type": "application/json"}
        data = {
            "model": self.model_name,
            "messages": [{"role": "user", "content": "hello"}],
            "max_tokens": 3, "temperature": 0.6, "top_p": 0.95, "stream": False,
        }

        for i in range(1, retry_count + 1):
            try:
                resp = requests.post(self.url, json=data, headers=headers, timeout=30)
                if resp.status_code == 200:
                    logger.info(f"健康检查通过 (第{i}次)")
                    return True
                else:
                    logger.warning(f"健康检查失败 (状态码 {resp.status_code}), {interval}s 后重试 ({i}/{retry_count})")
            except requests.RequestException as e:
                logger.warning(f"健康检查异常: {e}, {interval}s 后重试 ({i}/{retry_count})")
            time.sleep(interval)

        logger.error("健康检查超时失败")
        return False

    def get_container_logs(self):
        """获取容器日志，用于启动失败时记录"""
        if not self.container_name:
            return "未解析到容器名称，无法获取日志"

        prefer_sudo = ("sudo docker" in (self.container_cmd or "")) or ("sudo docker" in (self.stop_cmd or ""))
        base_cmd = f"docker logs {self.container_name} --tail 200"
        cmd_candidates = [f"sudo {base_cmd}", base_cmd] if prefer_sudo else [base_cmd, f"sudo {base_cmd}"]

        errors = []
        try:
            for cmd in cmd_candidates:
                logger.info(f"获取容器日志命令: {cmd}")
                result = self._run_cmd(cmd, timeout=30)
                stdout = (result.stdout or "").strip()
                stderr = (result.stderr or "").strip()
                if result.returncode == 0:
                    return stdout or "容器日志为空（命令执行成功但无输出）"
                errors.append(f"[{cmd}] rc={result.returncode}, err={stderr or stdout or '无输出'}")

            # 两种日志命令都失败时，补充容器状态便于定位
            ps_base = f"docker ps -a --filter name={self.container_name} --format '{{{{.Names}}}} {{{{.Status}}}}'"
            ps_candidates = [f"sudo {ps_base}", ps_base] if prefer_sudo else [ps_base, f"sudo {ps_base}"]
            status_output = ""
            for ps_cmd in ps_candidates:
                logger.info(f"获取容器状态命令: {ps_cmd}")
                ps_result = self._run_cmd(ps_cmd, timeout=15)
                if ps_result.returncode == 0:
                    status_output = (ps_result.stdout or "").strip()
                    if status_output:
                        break

            status_msg = f"; 容器状态: {status_output}" if status_output else ""
            return "获取容器日志失败: " + " | ".join(errors) + status_msg
        except Exception as e:
            return f"获取容器日志失败: {e}"


# ============================================================
# Benchmark (evalscope 压测)
# ============================================================

def run_single_benchmark(bench_config, context, batch_req):
    """执行单次压测，返回解析后的指标 dict 或 None"""
    input_tokens, output_tokens = context
    batch_size, request_count = batch_req
    logger.info(
        f"▶ 测试: in={input_tokens}, out={output_tokens}, "
        f"concurrency={batch_size}, requests={request_count}"
    )

    bench_args = Arguments(
        parallel=[batch_size],
        number=[request_count],
        model=bench_config["model_name"],
        url=bench_config["url"],
        api_key=bench_config.get("api_key") or None,
        tokenizer_path=bench_config["tokenizer_path"],
        api="openai",
        dataset=bench_config.get("dataset", "random"),
        min_tokens=output_tokens,
        max_tokens=output_tokens,
        min_prompt_length=input_tokens,
        max_prompt_length=input_tokens,
        prefix_length=0,
        debug=False,
        extra_args={"ignore_eos": True},
    )

    try:
        result = run_perf_benchmark(bench_args)
        logger.debug(f"Raw benchmark result: {result}")
        metrics = _extract_metrics(result)
        return _parse_metrics(metrics)
    except SystemExit as e:
        logger.error(f"基准测试异常退出: {e}")
        return None
    except ConnectionRefusedError as e:
        logger.error(f"无法连接到模型服务: {e}")
        return None
    except Exception as e:
        logger.error(f"基准测试执行失败: {e}")
        return None


# warmup 默认参数
WARMUP_DEFAULTS = {
    "context": (1024, 1024),
    "parallel": 5,
    "number": 10,
}


def run_warmup(bench_config):
    """预热: 支持配置上下文、并发数、请求数"""
    warmup_conf = bench_config.get("warmup", {})
    if isinstance(warmup_conf, bool):
        warmup_conf = {}
    # 解析 context
    ctx = warmup_conf.get("context", WARMUP_DEFAULTS["context"])
    if isinstance(ctx, str):
        ctx = ast.literal_eval(ctx)
    input_tokens, output_tokens = ctx
    parallel = int(warmup_conf.get("parallel", WARMUP_DEFAULTS["parallel"]))
    number = int(warmup_conf.get("number", WARMUP_DEFAULTS["number"]))

    logger.info(f"开始预热: in={input_tokens}, out={output_tokens}, "
                f"parallel={parallel}, number={number}")
    warmup_args = Arguments(
        parallel=[parallel], number=[number],
        model=bench_config["model_name"],
        url=bench_config["url"],
        api_key=bench_config.get("api_key") or None,
        tokenizer_path=bench_config["tokenizer_path"],
        api="openai",
        dataset=bench_config.get("dataset", "random"),
        min_tokens=output_tokens, max_tokens=output_tokens,
        min_prompt_length=input_tokens, max_prompt_length=input_tokens,
        prefix_length=0, debug=False,
        extra_args={"ignore_eos": True},
    )
    try:
        run_perf_benchmark(warmup_args)
        logger.info("预热完成")
        return True
    except Exception as e:
        logger.warning(f"预热失败，将继续正式测试: {e}")
        return False


def check_slo_criteria(result, criteria_parsed):
    """检查结果是否满足 SLO 准则"""
    if result is None:
        return False
    for metric, threshold in criteria_parsed.items():
        try:
            value = float(result[metric])
        except (KeyError, ValueError, TypeError):
            logger.warning(f"指标检查失败: {metric} 缺失或不可解析")
            return False
        if value > threshold:
            logger.warning(f"❌ 指标不达标: {metric}={value:.6f}s > 阈值 {threshold:.6f}s")
            return False
    return True


def _extract_metrics(result):
    """从 evalscope 返回值中提取 metrics dict"""
    if isinstance(result, list) and len(result) > 0:
        data = result[0]
    else:
        data = result
    if isinstance(data, dict):
        for v in data.values():
            if isinstance(v, dict) and "metrics" in v:
                return v["metrics"]
        return data
    return {}


def _parse_metrics(metrics):
    """将 evalscope 原始 metrics 转换为统一格式"""
    if not metrics:
        return None
    comment = ""
    failed = int(metrics.get("Failed requests", 0))
    total = metrics.get("Total requests", "N/A")
    if failed > 0:
        comment = f"失败请求数: {failed}/{total}"
    return {
        "ttft": metrics.get("Average time to first token (s)", "N/A"),
        "tpot": metrics.get("Average time per output token (s)", "N/A"),
        "throughput": metrics.get("Total token throughput (tok/s)", "N/A"),
        "itl": metrics.get("Average inter-token latency (s)", "N/A"),
        "duration": metrics.get("Time taken for tests (s)", "N/A"),
        "comment": comment,
    }


# ============================================================
# Excel 结果写入
# ============================================================

class Mode1ResultWriter:
    """模式1结果写入器：每个参数组合一个 sheet"""

    def __init__(self, result_file):
        self.result_file = result_file
        self.wb = Workbook()
        self._first_sheet = True

    def create_param_sheet(self, param_index, full_cmd, engine_name=None):
        title = f"{engine_name}-参数{param_index}" if engine_name else f"参数{param_index}"
        if self._first_sheet:
            ws = self.wb.active
            ws.title = title
            self._first_sheet = False
        else:
            ws = self.wb.create_sheet(title=title)
        headers = [
            "输入上下文长度", "输出上下文长度", "并发数", "请求数",
            "TTFT(s)", "TPOT(s)", "吞吐(tokens/s)", "ITL(s)", "持续时间(s)", "备注",
        ]
        ws.append(headers)
        ws._full_cmd = full_cmd
        return ws

    def append_result(self, ws, context, batch_req, result):
        failed_request = False
        if result is None:
            row = [context[0], context[1], batch_req[0], batch_req[1],
                   "N/A", "N/A", "N/A", "N/A", "N/A", "模型服务异常，未能获取结果"]
            failed_request = True
        else:
            row = [context[0], context[1], batch_req[0], batch_req[1],
                   result["ttft"], result["tpot"], result["throughput"],
                   result["itl"], result["duration"], result["comment"]]
            failed_request = bool(result.get("comment"))
        ws.append(row)
        if failed_request:
            red_font = Font(color="FFFF0000")
            row_idx = ws.max_row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).font = red_font

    def append_failure_marker(self, ws, context, reason):
        """在明细表中追加失败标注行（整行红色）"""
        row = [
            context[0], context[1], "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A", reason,
        ]
        ws.append(row)
        red_font = Font(color="FFFF0000")
        row_idx = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).font = red_font

    def append_failure_logs(self, ws, logs, title="失败日志"):
        """在参数 sheet 中追加失败日志"""
        logs_text = _sanitize_for_excel(str(logs or "").strip())
        if not logs_text:
            return
        max_len = 15000
        if len(logs_text) > max_len:
            logs_text = logs_text[:max_len] + "\n...(日志已截断)"

        ws.append([])
        ws.append([f"{title}:"])
        ws.append([logs_text])

        red_font = Font(color="FFFF0000")
        title_row = ws.max_row - 1
        content_row = ws.max_row
        ws.cell(row=title_row, column=1).font = red_font
        ws.cell(row=content_row, column=1).font = red_font

    def finalize_sheet(self, ws):
        full_cmd = getattr(ws, "_full_cmd", "")
        if full_cmd:
            ws.append([])
            ws.append(["完整启动命令:"])
            ws.append([full_cmd])

    def flush(self):
        """增量保存：每个 case 测试完后调用，立即落盘"""
        os.makedirs(os.path.dirname(self.result_file), exist_ok=True)
        self.wb.save(self.result_file)

    def save(self):
        os.makedirs(os.path.dirname(self.result_file), exist_ok=True)
        self.wb.save(self.result_file)
        logger.info(f"结果已保存: {self.result_file}")


class Mode2ResultWriter:
    """模式2结果写入器：每个引擎独立汇总 sheet + 引擎前缀详细 sheet"""

    def __init__(self, result_file, bench_only=False):
        self.result_file = result_file
        self.wb = Workbook()
        # 删除默认空 sheet
        self.wb.remove(self.wb.active)
        self.bench_only = bench_only
        # 按引擎名存储各引擎的状态
        self._engines = {}     # engine_name -> dict
        self.detail_sheets = []

    def init_engine(self, engine_name, exclude_params=None):
        """为一个引擎初始化汇总 sheet 和相关数据结构"""
        summary_ws = self.wb.create_sheet(title=f"汇总-{engine_name}")
        arranged_ws = self.wb.create_sheet(title=f"整理-汇总-{engine_name}")
        self._engines[engine_name] = {
            "summary_ws": summary_ws,
            "arranged_ws": arranged_ws,
            "summary_rows": [],
            "commands_params": [],
            "exclude_params": exclude_params or [],
        }

    def register_command(self, engine_name, model_cmd):
        params = parse_model_cmd_args(model_cmd)
        self._engines[engine_name]["commands_params"].append(params)

    def create_detail_sheet(self, engine_name, param_index, full_cmd):
        ws = self.wb.create_sheet(title=f"{engine_name}-参数{param_index}")
        headers = [
            "输入上下文长度", "输出上下文长度", "并发数", "请求数",
            "TTFT(s)", "TPOT(s)", "吞吐(tokens/s)", "ITL(s)", "持续时间(s)", "备注",
        ]
        ws.append(headers)
        ws._full_cmd = full_cmd
        ws._effective_cmds_list = []   # [(context, effective_full_cmd), ...]
        self.detail_sheets.append(ws)
        return ws

    def append_detail_result(self, ws, context, batch_req, result):
        failed_request = False
        if result is None:
            row = [context[0], context[1], batch_req[0], batch_req[1],
                   "N/A", "N/A", "N/A", "N/A", "N/A", "模型服务异常，未能获取结果"]
            failed_request = True
        else:
            row = [context[0], context[1], batch_req[0], batch_req[1],
                   result["ttft"], result["tpot"], result["throughput"],
                   result["itl"], result["duration"], result["comment"]]
            failed_request = bool(result.get("comment"))
        ws.append(row)
        if failed_request:
            red_font = Font(color="FFFF0000")
            row_idx = ws.max_row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).font = red_font

    def append_failure_marker(self, ws, context, reason):
        """在明细表中追加失败标注行（整行红色）"""
        row = [
            context[0], context[1], "N/A", "N/A",
            "N/A", "N/A", "N/A", "N/A", "N/A", reason,
        ]
        ws.append(row)
        red_font = Font(color="FFFF0000")
        row_idx = ws.max_row
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).font = red_font

    def append_failure_logs(self, ws, logs, title="失败日志"):
        """在参数 sheet 中追加失败日志"""
        logs_text = _sanitize_for_excel(str(logs or "").strip())
        if not logs_text:
            return
        max_len = 15000
        if len(logs_text) > max_len:
            logs_text = logs_text[:max_len] + "\n...(日志已截断)"

        ws.append([])
        ws.append([f"{title}:"])
        ws.append([logs_text])

        red_font = Font(color="FFFF0000")
        title_row = ws.max_row - 1
        content_row = ws.max_row
        ws.cell(row=title_row, column=1).font = red_font
        ws.cell(row=content_row, column=1).font = red_font

    def highlight_best_rows(self, ws, context_best_map, contexts):
        bold_font = Font(bold=True)
        green_fill = PatternFill(fill_type="solid", start_color="FF92D050", end_color="FF92D050")
        yellow_fill = PatternFill(fill_type="solid", start_color="FFFFFF00", end_color="FFFFFF00")
        for item in contexts:
            ctx = item["context"]
            ctx_data = context_best_map.get(tuple(ctx))
            if not ctx_data:
                continue
            best = ctx_data.get("best") if isinstance(ctx_data, dict) else ctx_data
            min_tested = ctx_data.get("min_tested") if isinstance(ctx_data, dict) else None
            if best:
                target_concurrency = best["batch_req"][0]
                row_fill = green_fill
            elif min_tested:
                target_concurrency = min_tested["batch_req"][0]
                row_fill = yellow_fill
            else:
                continue
            for row_idx in range(2, ws.max_row + 1):
                if (ws.cell(row=row_idx, column=1).value == ctx[0]
                        and ws.cell(row=row_idx, column=2).value == ctx[1]
                        and ws.cell(row=row_idx, column=3).value == target_concurrency):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = bold_font
                        cell.fill = row_fill
                    break

    def add_summary_row(self, engine_name, model_name, image_name, param_index, context, best_result,
                        failure_reason=None, effective_model_cmd=None, slo_no_pass=False):
        self._engines[engine_name]["summary_rows"].append({
            "model_name": model_name, "image_name": image_name,
            "param_index": param_index,
            "input_ctx": context[0] if context else "",
            "output_ctx": context[1] if context else "",
            "best": best_result,
            "failure_reason": failure_reason,
            "effective_model_cmd": effective_model_cmd,
            "slo_no_pass": slo_no_pass,
        })

    def finalize_detail_sheet(self, ws):
        if getattr(ws, "_finalized", False):
            return
        full_cmd = getattr(ws, "_full_cmd", "")
        effective_cmds_list = getattr(ws, "_effective_cmds_list", [])

        # 去重：收集不同的有效命令
        distinct_cmds = []
        seen_cmds = set()
        for ctx, cmd in effective_cmds_list:
            if cmd not in seen_cmds:
                seen_cmds.add(cmd)
                distinct_cmds.append((ctx, cmd))

        if len(distinct_cmds) > 1:
            ws.append([])
            ws.append(["各上下文启动命令:"])
            for ctx, cmd in effective_cmds_list:
                ws.append([f"上下文 {ctx}: {cmd}"])
        elif distinct_cmds:
            ws.append([])
            ws.append(["完整启动命令:"])
            ws.append([distinct_cmds[0][1]])
        elif full_cmd:
            ws.append([])
            ws.append(["完整启动命令:"])
            ws.append([full_cmd])
        ws._finalized = True

    def write_all_summaries(self):
        metric_headers = ["并发", "请求数", "TTFT(s)", "TPOT(s)", "吞吐(tokens/s)"]

        def _apply_row_style(ws, row, style):
            ws.append(row)
            if style == "red":
                font = Font(color="FFFF0000")
            elif style == "yellow":
                font = Font(color="FFFFCC00")
            else:
                return
            row_idx = ws.max_row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).font = font

        for engine_name, engine_data in self._engines.items():
            ws = engine_data["summary_ws"]
            arranged_ws = engine_data["arranged_ws"]
            # 支持重复调用：每次先清空后重写，避免重复行
            if ws.max_row > 0:
                ws.delete_rows(1, ws.max_row)
            if arranged_ws.max_row > 0:
                arranged_ws.delete_rows(1, arranged_ws.max_row)
            summary_rows = engine_data["summary_rows"]
            commands_params = engine_data["commands_params"]
            exclude_params = engine_data["exclude_params"]

            normalized_rows = []

            if self.bench_only:
                fixed_headers = ["输入上下文", "输出上下文"]
                ws.append(fixed_headers + metric_headers)
                arranged_headers = ["输入上下文", "输出上下文", "参数序号"] + metric_headers
                arranged_ws.append(arranged_headers)

                for row_data in summary_rows:
                    best = row_data["best"]
                    param_index = row_data["param_index"]
                    fixed = [row_data["input_ctx"], row_data["output_ctx"]]
                    failed_request = False
                    slo_no_pass = row_data.get("slo_no_pass", False)
                    failure_reason = row_data.get("failure_reason")
                    if best:
                        br = best["batch_req"]
                        r = best.get("result") or {}
                        metrics = [br[0], br[1], r.get("ttft", "N/A"), r.get("tpot", "N/A"), r.get("throughput", "N/A")]
                        failed_request = bool(r.get("comment"))
                    else:
                        metrics = ["N/A", "N/A", "N/A", "N/A", "N/A"]
                    if failure_reason:
                        failed_request = True
                    row_style = "red" if failed_request else ("yellow" if slo_no_pass else None)
                    _apply_row_style(ws, fixed + metrics, row_style)
                    normalized_rows.append({
                        "input_ctx": row_data["input_ctx"],
                        "output_ctx": row_data["output_ctx"],
                        "param_index": param_index,
                        "arranged_tail": metrics,
                        "failed_request": failed_request,
                        "slo_no_pass": slo_no_pass,
                    })
            else:
                # 预计算每行的有效参数（优先使用 effective_model_cmd）
                row_params_list = []
                for row_data in summary_rows:
                    effective_cmd = row_data.get("effective_model_cmd")
                    if effective_cmd:
                        row_data["_params"] = parse_model_cmd_args(effective_cmd)
                    else:
                        idx = row_data["param_index"]
                        row_data["_params"] = commands_params[idx - 1] if idx <= len(commands_params) else {}
                    row_params_list.append(row_data["_params"])

                all_param_keys = collect_all_param_keys(
                    commands_params + row_params_list, exclude=exclude_params)
                fixed_headers = ["模型名称", "参数序号", "推理引擎(镜像)", "输入上下文", "输出上下文"]
                param_headers = list(all_param_keys)
                ws.append(fixed_headers + param_headers + metric_headers)
                arranged_headers = [
                    "输入上下文", "输出上下文", "参数序号", "模型名称", "推理引擎(镜像)"
                ] + param_headers + metric_headers
                arranged_ws.append(arranged_headers)

                for row_data in summary_rows:
                    params = row_data["_params"]
                    idx = row_data["param_index"]
                    best = row_data["best"]
                    failed_request = False
                    slo_no_pass = row_data.get("slo_no_pass", False)
                    failure_reason = row_data.get("failure_reason")

                    fixed = [row_data["model_name"], idx, row_data["image_name"],
                             row_data["input_ctx"], row_data["output_ctx"]]
                    param_vals = [params.get(k, "") for k in all_param_keys]

                    if best:
                        br = best["batch_req"]
                        r = best.get("result") or {}
                        metrics = [br[0], br[1], r.get("ttft", "N/A"), r.get("tpot", "N/A"), r.get("throughput", "N/A")]
                        failed_request = bool(r.get("comment"))
                    else:
                        metrics = ["N/A", "N/A", "N/A", "N/A", "N/A"]
                    if failure_reason:
                        failed_request = True

                    row_style = "red" if failed_request else ("yellow" if slo_no_pass else None)
                    _apply_row_style(ws, fixed + param_vals + metrics, row_style)
                    normalized_rows.append({
                        "input_ctx": row_data["input_ctx"],
                        "output_ctx": row_data["output_ctx"],
                        "param_index": idx,
                        "arranged_tail": [row_data["model_name"], row_data["image_name"]] + param_vals + metrics,
                        "failed_request": failed_request,
                        "slo_no_pass": slo_no_pass,
                    })

            normalized_rows.sort(key=lambda r: (r["input_ctx"], r["output_ctx"], r["param_index"]))
            for row in normalized_rows:
                arranged_row = [row["input_ctx"], row["output_ctx"], row["param_index"]] + row["arranged_tail"]
                row_style = "red" if row["failed_request"] else ("yellow" if row.get("slo_no_pass", False) else None)
                _apply_row_style(arranged_ws, arranged_row, row_style)

    def flush(self):
        """增量保存：每个 case 测试完后调用，立即落盘"""
        os.makedirs(os.path.dirname(self.result_file), exist_ok=True)
        self.wb.save(self.result_file)

    def save(self):
        self.write_all_summaries()
        for ws in self.detail_sheets:
            self.finalize_detail_sheet(ws)
        os.makedirs(os.path.dirname(self.result_file), exist_ok=True)
        self.wb.save(self.result_file)
        logger.info(f"结果已保存: {self.result_file}")


# ============================================================
# 主流程
# ============================================================

def run_mode1(bench_config, model_config):
    """模式1: 固定测试case"""
    cases = bench_config["test_cases"]["cases"]
    if not cases:
        logger.error("未生成有效测试用例")
        return

    bench_only = bench_config.get("bench_only", False)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    result_file = f"{bench_config['result_dir']}/{bench_config['result_file_prefix']}-{timestamp}.xlsx"
    writer = Mode1ResultWriter(result_file)
    engines = model_config["engines"]

    if bench_only:
        logger.info("========== 纯测试模式（bench_only） ==========")

    for eng_idx, engine in enumerate(engines, 1):
        engine_name = engine["engine_name"]
        manager = ModelManager(engine, bench_config)
        commands = engine["commands"]
        logger.info(f"====== 推理引擎 {eng_idx}/{len(engines)}: {engine_name} ======")

        for cmd_idx, cmd_item in enumerate(commands, 1):
            model_cmd = cmd_item["cmd"]
            full_cmd = manager.build_full_cmd(model_cmd)
            logger.info(f"========== {engine_name} 参数组{cmd_idx}/{len(commands)} ==========")

            if not bench_only:
                # 启动模型
                success, logs = manager.start_model(model_cmd)
                if not success:
                    logger.error(f"{engine_name} 参数组{cmd_idx} 模型启动失败，跳过\n日志:\n{logs}")
                    ws = writer.create_param_sheet(cmd_idx, full_cmd, engine_name)
                    for ctx, br in cases:
                        writer.append_result(ws, ctx, br, None)
                    writer.append_failure_logs(ws, logs, title="启动失败日志")
                    writer.finalize_sheet(ws)
                    manager.stop_model()
                    continue

                # 健康检查
                if not manager.health_check():
                    container_logs = manager.get_container_logs()
                    logger.error(f"{engine_name} 参数组{cmd_idx} 健康检查失败\n容器日志:\n{container_logs}")
                    ws = writer.create_param_sheet(cmd_idx, full_cmd, engine_name)
                    for ctx, br in cases:
                        writer.append_result(ws, ctx, br, None)
                    writer.append_failure_logs(ws, container_logs, title="健康检查失败日志")
                    writer.finalize_sheet(ws)
                    manager.stop_model()
                    continue

            elif bench_config.get("bench_only_healthcheck", False):
                # bench_only 模式下可选健康检查
                logger.info("执行 bench_only 健康检查...")
                if not manager.health_check():
                    logger.error(f"{engine_name} 参数组{cmd_idx} bench_only 健康检查失败，跳过")
                    ws = writer.create_param_sheet(cmd_idx, full_cmd, engine_name)
                    for ctx, br in cases:
                        writer.append_result(ws, ctx, br, None)
                    writer.append_failure_logs(ws, "健康检查失败（bench_only 模式）", title="健康检查失败")
                    writer.finalize_sheet(ws)
                    continue

            # 预热
            warmup_conf = bench_config.get("warmup", {})
            if (isinstance(warmup_conf, dict) and warmup_conf.get("enable", False)) or warmup_conf is True:
                run_warmup(bench_config)

            # 执行测试
            ws = writer.create_param_sheet(cmd_idx, full_cmd, engine_name)
            for i, (ctx, br) in enumerate(cases, 1):
                logger.info(f"测试进度: {i}/{len(cases)}")
                result = run_single_benchmark(bench_config, ctx, br)
                logger.info(f"测试结果: {result}")
                writer.append_result(ws, ctx, br, result)
                writer.flush()
            writer.finalize_sheet(ws)

            if not bench_only:
                manager.stop_model()

    writer.save()
    logger.info("模式1 全部测试完成!")


def run_mode2(bench_config, model_config):
    """模式2: SLO最大并发探测"""
    slo = bench_config["slo"]
    contexts = slo["contexts"]
    criteria_parsed = slo["criteria_parsed"]
    default_init = slo.get("init_concurrent", 2)
    default_max = slo.get("max_concurrent", None)
    multiplier = slo.get("request_multiplier", 1)
    search_method = slo.get("search_method", "binary")
    stop_n = int(slo.get("stop_n", 2))

    if not contexts:
        logger.error("未配置 SLO context")
        return

    bench_only = bench_config.get("bench_only", False)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    result_file = f"{bench_config['result_dir']}/{bench_config['result_file_prefix']}-slo-{timestamp}.xlsx"
    writer = Mode2ResultWriter(result_file, bench_only=bench_only)
    engines = model_config["engines"]

    if bench_only:
        logger.info("========== 纯测试模式（bench_only） ==========")

    for eng_idx, engine in enumerate(engines, 1):
        engine_name = engine["engine_name"]
        manager = ModelManager(engine, bench_config)
        commands = engine["commands"]
        image_name = engine.get("image_name", "")
        writer.init_engine(engine_name, exclude_params=engine.get("exclude_params", []))
        logger.info(f"====== 推理引擎 {eng_idx}/{len(engines)}: {engine_name} ======")

        for cmd_idx, cmd_item in enumerate(commands, 1):
            model_cmd = cmd_item["cmd"]
            full_cmd = manager.build_full_cmd(model_cmd)
            writer.register_command(engine_name, model_cmd)
            logger.info(f"========== {engine_name} 参数组{cmd_idx}/{len(commands)} ==========")

            detail_ws = writer.create_detail_sheet(engine_name, cmd_idx, full_cmd)
            context_best_map = {}

            if bench_only:
                # bench_only 模式：无需启停模型，直接测试
                if bench_config.get("bench_only_healthcheck", False):
                    logger.info("执行 bench_only 健康检查...")
                    if not manager.health_check():
                        logger.error(f"{engine_name} 参数组{cmd_idx} bench_only 健康检查失败，跳过")
                        writer.finalize_detail_sheet(detail_ws)
                        writer.write_all_summaries()
                        writer.flush()
                        continue

                warmup_conf = bench_config.get("warmup", {})
                if (isinstance(warmup_conf, dict) and warmup_conf.get("enable", False)) or warmup_conf is True:
                    run_warmup(bench_config)

                for ctx_idx, item in enumerate(contexts, 1):
                    context = item["context"]
                    init_concurrent = item.get("init_concurrent") or default_init
                    max_concurrent = item.get("max_concurrent") or default_max
                    logger.info(f"--- 上下文 {context} ({ctx_idx}/{len(contexts)}) ---")

                    best, min_tested = _search_max_concurrency(
                        bench_config, context, init_concurrent, multiplier,
                        criteria_parsed, search_method, stop_n, detail_ws, writer,
                        max_concurrent=max_concurrent
                    )
                    context_best_map[tuple(context)] = {"best": best, "min_tested": min_tested}
                    slo_no_pass = best is None and min_tested is not None
                    writer.flush()
                    writer.add_summary_row(
                        engine_name, bench_config["model_name"], image_name,
                        cmd_idx, context, best if best is not None else min_tested,
                        slo_no_pass=slo_no_pass)
            else:
                # 非 bench_only：支持 per-context override_args，按需重启模型
                current_running_cmd = None

                for ctx_idx, item in enumerate(contexts, 1):
                    context = item["context"]
                    init_concurrent = item.get("init_concurrent") or default_init
                    max_concurrent = item.get("max_concurrent") or default_max
                    override_args = item.get("override_args", [])
                    engine_override_args = item.get("engine_override_args", {})
                    # 合并通用覆盖 + 引擎专属覆盖
                    effective_overrides = list(override_args) + list(engine_override_args.get(engine_name, []))

                    # 计算有效命令（应用 override_args）
                    effective_cmd = apply_override_args(model_cmd, effective_overrides)
                    effective_full_cmd = manager.build_full_cmd(effective_cmd)

                    logger.info(f"--- 上下文 {context} ({ctx_idx}/{len(contexts)}) ---")
                    if effective_overrides:
                        logger.info(f"上下文参数覆盖(通用): {override_args}")
                        if engine_override_args.get(engine_name):
                            logger.info(f"上下文参数覆盖({engine_name}): {engine_override_args[engine_name]}")
                        logger.info(f"有效启动命令: {effective_cmd}")

                    # 记录有效命令到明细表（用于 finalize_detail_sheet）
                    detail_ws._effective_cmds_list.append((list(context), effective_full_cmd))

                    # 判断是否需要（重新）启动模型
                    need_restart = False
                    if current_running_cmd is None:
                        need_restart = True
                    elif not _cmds_are_equivalent(effective_cmd, current_running_cmd):
                        logger.info(
                            f"上下文 {context} 的参数与当前运行参数不同，需要重启模型")
                        manager.stop_model()
                        need_restart = True

                    if need_restart:
                        success, logs = manager.start_model(effective_cmd)
                        if not success:
                            logger.error(
                                f"{engine_name} 参数组{cmd_idx} 上下文{context} "
                                f"模型启动失败\n日志:\n{logs}")
                            failure_reason = "模型启动失败（参数或环境异常）"
                            writer.append_failure_marker(
                                detail_ws, context, failure_reason)
                            writer.add_summary_row(
                                engine_name, bench_config["model_name"],
                                image_name, cmd_idx, context, None,
                                failure_reason=failure_reason,
                                effective_model_cmd=effective_cmd)
                            writer.append_failure_logs(
                                detail_ws, logs, title="启动失败日志")
                            writer.write_all_summaries()
                            writer.flush()
                            current_running_cmd = None
                            continue

                        if not manager.health_check():
                            container_logs = manager.get_container_logs()
                            logger.error(
                                f"{engine_name} 参数组{cmd_idx} 上下文{context} "
                                f"健康检查失败\n容器日志:\n{container_logs}")
                            failure_reason = "健康检查失败（服务未就绪）"
                            writer.append_failure_marker(
                                detail_ws, context, failure_reason)
                            writer.add_summary_row(
                                engine_name, bench_config["model_name"],
                                image_name, cmd_idx, context, None,
                                failure_reason=failure_reason,
                                effective_model_cmd=effective_cmd)
                            writer.append_failure_logs(
                                detail_ws, container_logs,
                                title="健康检查失败日志")
                            writer.write_all_summaries()
                            writer.flush()
                            manager.stop_model()
                            current_running_cmd = None
                            continue

                        current_running_cmd = effective_cmd

                        # 预热（仅在模型（重新）启动后）
                        warmup_conf = bench_config.get("warmup", {})
                        if (isinstance(warmup_conf, dict) and warmup_conf.get("enable", False)) or warmup_conf is True:
                            run_warmup(bench_config)

                    # SLO 并发探测
                    best, min_tested = _search_max_concurrency(
                        bench_config, context, init_concurrent, multiplier,
                        criteria_parsed, search_method, stop_n, detail_ws, writer,
                        max_concurrent=max_concurrent
                    )
                    context_best_map[tuple(context)] = {"best": best, "min_tested": min_tested}
                    slo_no_pass = best is None and min_tested is not None
                    writer.flush()
                    writer.add_summary_row(
                        engine_name, bench_config["model_name"], image_name,
                        cmd_idx, context, best if best is not None else min_tested,
                        slo_no_pass=slo_no_pass,
                        effective_model_cmd=effective_cmd)

                # 停止当前运行的模型（如果有）
                if current_running_cmd is not None:
                    manager.stop_model()

            writer.highlight_best_rows(detail_ws, context_best_map, contexts)
            _log_context_summary(contexts, context_best_map)
            writer.finalize_detail_sheet(detail_ws)
            writer.write_all_summaries()
            writer.flush()

    writer.save()
    logger.info("模式2 全部测试完成!")


def _log_context_summary(contexts, context_best_map):
    """输出各上下文最高达标并发汇总日志"""
    logger.info("========== 各上下文最高达标并发 ==========")
    for item in contexts:
        ctx = item["context"]
        ctx_data = context_best_map.get(tuple(ctx))
        best = ctx_data.get("best") if isinstance(ctx_data, dict) else ctx_data
        min_tested = ctx_data.get("min_tested") if isinstance(ctx_data, dict) else None
        if best:
            logger.info(f"  ✅ {ctx} -> 并发={best['batch_req'][0]}, 请求数={best['batch_req'][1]}")
        elif min_tested:
            logger.info(f"  ❌ {ctx} -> 未找到达标并发（最小测试并发={min_tested['batch_req'][0]}，结果已标记黄色）")
        else:
            logger.info(f"  ❌ {ctx} -> 未找到达标并发")


def _search_max_concurrency(bench_config, context, init_concurrent, multiplier,
                            criteria_parsed, search_method, stop_n,
                            detail_ws, writer, max_concurrent=None):
    """搜索满足 SLO 的最大并发数，返回 best dict 或 None"""
    concurrent = init_concurrent
    if max_concurrent is not None:
        max_concurrent = int(max_concurrent)
        concurrent = min(concurrent, max_concurrent)
        logger.info(f"最大并发限制: {max_concurrent}")
    lower_bound = 0
    upper_bound = None
    best = None
    tested = set()
    all_tested = []

    while True:
        # 限制并发不超过 max_concurrent
        if max_concurrent is not None and concurrent > max_concurrent:
            logger.info(f"并发 {concurrent} 超过最大限制 {max_concurrent}，截断为 {max_concurrent}")
            concurrent = max_concurrent
        if concurrent in tested:
            logger.info("搜索结束，当前并发已测试过")
            break
        tested.add(concurrent)

        req_count = int(concurrent * multiplier)
        batch_req = (concurrent, req_count)

        result = run_single_benchmark(bench_config, context, batch_req)
        all_tested.append((concurrent, batch_req, result))
        writer.append_detail_result(detail_ws, context, batch_req, result)
        # 每次 evalscope 执行完都同步刷新汇总并立即落盘
        writer.write_all_summaries()
        writer.flush()
        is_valid = check_slo_criteria(result, criteria_parsed)
        logger.info(f"结果: {result}")
        logger.info(f"SLO检查: {'✅ 达标' if is_valid else '❌ 未达标'}")

        if search_method == "stopN":
            concurrent, lower_bound, upper_bound, best, should_break = _stopn_step(
                concurrent, is_valid, lower_bound, upper_bound, stop_n,
                best, batch_req, result, tested)
        else:
            concurrent, lower_bound, upper_bound, best, should_break = _binary_step(
                concurrent, is_valid, lower_bound, upper_bound, init_concurrent,
                best, batch_req, result)

        if should_break:
            break

    min_tested = None
    if best is None and all_tested:
        min_entry = min(all_tested, key=lambda x: x[0])
        min_tested = {"batch_req": min_entry[1], "result": min_entry[2]}
    return best, min_tested


def _stopn_step(concurrent, is_valid, lower_bound, upper_bound, stop_n,
                best, batch_req, result, tested):
    should_break = False
    if is_valid:
        if best is None or concurrent > best["batch_req"][0]:
            best = {"batch_req": batch_req, "result": result}
        lower_bound = max(lower_bound, concurrent)
        if upper_bound is None:
            next_c = concurrent + stop_n
            logger.info(f"✅ 达标，并发 +{stop_n}: {concurrent} -> {next_c}")
        else:
            next_c = (lower_bound + upper_bound) // 2
            logger.info(f"✅ 达标，区间 [{lower_bound}, {upper_bound}] 二分: {next_c}")
    else:
        upper_bound = concurrent if upper_bound is None else min(upper_bound, concurrent)
        if lower_bound == 0:
            next_c = max(1, concurrent - stop_n)
            logger.info(f"❌ 未达标，并发 -{stop_n}: {concurrent} -> {next_c}")
            if next_c == concurrent or (next_c == 1 and 1 in tested):
                should_break = True
        else:
            next_c = (lower_bound + upper_bound) // 2
            logger.info(f"❌ 未达标，区间 [{lower_bound}, {upper_bound}] 二分: {next_c}")

    if not should_break and next_c in tested:
        logger.info("搜索结束，下一并发已测试过")
        should_break = True

    return next_c, lower_bound, upper_bound, best, should_break


def _binary_step(concurrent, is_valid, lower_bound, upper_bound, init_concurrent,
                 best, batch_req, result):
    should_break = False
    if is_valid:
        if best is None or concurrent > best["batch_req"][0]:
            best = {"batch_req": batch_req, "result": result}
        lower_bound = concurrent
        if upper_bound is None:
            next_c = concurrent * 2
            logger.info(f"✅ 达标，翻倍: {concurrent} -> {next_c}")
        else:
            next_c = (lower_bound + upper_bound) // 2
            logger.info(f"✅ 达标，区间 [{lower_bound}, {upper_bound}] 二分: {next_c}")
    else:
        upper_bound = concurrent
        next_c = (lower_bound + upper_bound) // 2
        logger.info(f"❌ 未达标，区间 [{lower_bound}, {upper_bound}] 二分: {next_c}")

    if next_c == concurrent or next_c <= lower_bound:
        logger.info("搜索结束，最大有效并发已确认")
        should_break = True

    return next_c, lower_bound, upper_bound, best, should_break


# ============================================================
# 入口
# ============================================================

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parser = argparse.ArgumentParser(description="LLM 性能评测工具")
    parser.add_argument("-b", "--bench-config",
                        default=os.path.join(script_dir, "bench-conf.yaml"),
                        help="压测配置文件路径 (默认: 脚本同目录 bench-conf.yaml)")
    parser.add_argument("-m", "--model-config",
                        default=os.path.join(script_dir, "model-conf.yaml"),
                        help="模型启动配置文件路径 (默认: 脚本同目录 model-conf.yaml)")
    args = parser.parse_args()

    try:
        bench_config = load_bench_config(args.bench_config)
        mode = bench_config.get("mode", 1)
        bench_only = bench_config.get("bench_only", False)

        model_config = None
        model_conf_path = args.model_config

        if bench_only:
            logger.info("bench_only=true，跳过 model-conf 加载，使用 bench 配置运行")
            model_config = {
                "engines": [{
                    "engine_name": "bench-only",
                    "container_cmd": "",
                    "stop_cmd": "",
                    "ssh_cmd": None,
                    "container_name": None,
                    "image_name": "",
                    "exclude_params": [],
                    "commands": [{"cmd": "bench-only"}],
                }]
            }
        elif mode in (1, 2):
            if model_conf_path and os.path.exists(model_conf_path):
                model_config = load_model_config(model_conf_path)
            else:
                raise ValueError(f"mode={mode} 且 bench_only=false 时需要可用的模型配置文件: {model_conf_path}")
        else:
            logger.error(f"不支持的模式: {mode}")
            sys.exit(1)

        mode_desc = {1: "固定测试case", 2: "SLO最大并发探测"}.get(mode, "未知")
        engines = model_config["engines"]
        engine_names = [e["engine_name"] for e in engines]
        total_cmds = sum(len(e["commands"]) for e in engines)
        logger.info(f"测试模式: {mode} ({mode_desc}){' | 纯测试模式' if bench_only else ''}")
        logger.info(f"推理引擎: {', '.join(engine_names)} | 总命令组数: {total_cmds}")

        if mode == 1:
            run_mode1(bench_config, model_config)
        elif mode == 2:
            run_mode2(bench_config, model_config)
    except KeyboardInterrupt:
        logger.info("用户中断执行")
    except Exception as e:
        logger.error(f"程序异常退出: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
