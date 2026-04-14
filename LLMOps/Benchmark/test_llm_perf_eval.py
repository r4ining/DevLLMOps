#!/usr/bin/env python3
"""llm-perf-eval.py 离线测试脚本

无需真实模型服务即可验证脚本全部功能：
  - 工具函数（parse_time, parse_model_cmd_args, apply_override_args 等）
  - 配置加载（bench / model YAML 解析与默认值填充）
  - ModelManager 生命周期（mock subprocess / requests）
  - Benchmark 指标提取与 SLO 判定
  - 二分 / stopN 搜索算法
  - Excel 写入器（Mode1 / Mode2）
  - 端到端流程（run_mode1 / run_mode2，mock evalscope）

用法:
    python test_llm_perf_eval.py           # 运行全部测试
    python test_llm_perf_eval.py -v        # 详细输出
    python test_llm_perf_eval.py -k test_parse_time  # 只跑某个测试
"""
import ast
import os
import sys
import shutil
import tempfile
import textwrap
import unittest
from unittest import mock
from unittest.mock import patch, MagicMock

import yaml

# ---- 在导入主模块前 mock 掉 evalscope，避免依赖 ----
mock_evalscope_perf = MagicMock()
mock_evalscope_args = MagicMock()
sys.modules.setdefault("evalscope", MagicMock())
sys.modules.setdefault("evalscope.perf", MagicMock())
sys.modules.setdefault("evalscope.perf.main", mock_evalscope_perf)
sys.modules.setdefault("evalscope.perf.arguments", mock_evalscope_args)

# 让 Arguments 是一个普通类，可以接收任意关键字参数
class _FakeArguments:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)

mock_evalscope_args.Arguments = _FakeArguments

# 导入被测模块（文件名含连字符，需用 importlib 导入）
import importlib
import importlib.util

_spec = importlib.util.spec_from_file_location(
    "llm_perf_eval",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "llm-perf-eval.py"),
)
lpe = importlib.util.module_from_spec(_spec)
sys.modules["llm_perf_eval"] = lpe
_spec.loader.exec_module(lpe)

# ============================================================
# 辅助：生成临时 YAML 配置
# ============================================================

def _write_yaml(path, data):
    with open(path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, allow_unicode=True)


def _make_bench_yaml(tmp_dir, overrides=None):
    """生成最小可用的 bench 配置并返回文件路径"""
    conf = {
        "mode": 1,
        "model_name": "test-model",
        "tokenizer_path": "/fake/tokenizer",
        "url": "http://localhost:8000/v1/chat/completions",
        "test_cases": {
            "combination_mode": 1,
            "context": ["(512, 256)"],
            "batch_request": ["(1, 5)"],
        },
    }
    if overrides:
        conf.update(overrides)
    path = os.path.join(tmp_dir, "bench.yaml")
    _write_yaml(path, conf)
    return path


def _make_model_yaml(tmp_dir, overrides=None):
    """生成最小可用的 model 配置并返回文件路径"""
    conf = {
        "engines": [{
            "engine": "TestEngine",
            "container_cmd": "docker run -itd --name test-model-server test-image:latest",
            "stop_cmd": "docker stop test-model-server && docker rm -f test-model-server",
            "commands": [
                {"cmd": "python serve.py --tp-size 4 --host 0.0.0.0 --port 8000"},
            ],
        }],
    }
    if overrides:
        conf.update(overrides)
    path = os.path.join(tmp_dir, "model.yaml")
    _write_yaml(path, conf)
    return path


def _fake_benchmark_result(ttft=0.05, tpot=0.02, throughput=500, itl=0.02,
                           duration=10, failed=0, total=20):
    """构造 evalscope 返回值"""
    metrics = {
        "Average time to first token (s)": ttft,
        "Average time per output token (s)": tpot,
        "Total token throughput (tok/s)": throughput,
        "Average inter-token latency (s)": itl,
        "Time taken for tests (s)": duration,
        "Failed requests": failed,
        "Total requests": total,
    }
    return [{"dummy_key": {"metrics": metrics}}]


# ============================================================
# 测试: 工具函数
# ============================================================

class TestParseTime(unittest.TestCase):
    def test_seconds_string(self):
        self.assertAlmostEqual(lpe.parse_time("0.5s"), 0.5)

    def test_milliseconds_string(self):
        self.assertAlmostEqual(lpe.parse_time("200ms"), 0.2)

    def test_plain_number_string(self):
        self.assertAlmostEqual(lpe.parse_time("1.5"), 1.5)

    def test_numeric_input(self):
        self.assertAlmostEqual(lpe.parse_time(3), 3.0)

    def test_whitespace(self):
        self.assertAlmostEqual(lpe.parse_time("  100ms  "), 0.1)


class TestParseModelCmdArgs(unittest.TestCase):
    def test_simple_args(self):
        result = lpe.parse_model_cmd_args("python serve.py --tp-size 4 --host 0.0.0.0")
        self.assertEqual(result["tp-size"], "4")
        self.assertEqual(result["host"], "0.0.0.0")

    def test_bool_flag(self):
        result = lpe.parse_model_cmd_args("app --verbose --flag")
        self.assertEqual(result["verbose"], "true")
        self.assertEqual(result["flag"], "true")

    def test_negative_number_not_treated_as_key(self):
        result = lpe.parse_model_cmd_args("app --threshold -1")
        self.assertEqual(result["threshold"], "-1")

    def test_multiline_cmd(self):
        cmd = "python serve.py \\\n--tp-size 4 \\\n--port 8000"
        result = lpe.parse_model_cmd_args(cmd)
        self.assertEqual(result["tp-size"], "4")
        self.assertEqual(result["port"], "8000")


class TestApplyOverrideArgs(unittest.TestCase):
    def test_add_new_arg(self):
        result = lpe.apply_override_args("app --a 1", ["--b 2"])
        self.assertIn("--b", result)
        self.assertIn("2", result)

    def test_override_existing(self):
        result = lpe.apply_override_args("app --a 1 --b old", ["--b new"])
        params = lpe.parse_model_cmd_args(result)
        self.assertEqual(params["b"], "new")

    def test_delete_arg(self):
        result = lpe.apply_override_args("app --a 1 --b 2", ["!--b"])
        params = lpe.parse_model_cmd_args(result)
        self.assertNotIn("b", params)
        self.assertEqual(params["a"], "1")

    def test_empty_override(self):
        original = "app --a 1"
        self.assertEqual(lpe.apply_override_args(original, []), original)
        self.assertEqual(lpe.apply_override_args(original, None), original)

    def test_add_bool_flag(self):
        result = lpe.apply_override_args("app --a 1", ["--verbose"])
        params = lpe.parse_model_cmd_args(result)
        self.assertEqual(params["verbose"], "true")


class TestCmdsAreEquivalent(unittest.TestCase):
    def test_same_args_different_order(self):
        self.assertTrue(lpe._cmds_are_equivalent(
            "app --a 1 --b 2", "app --b 2 --a 1"))

    def test_different_args(self):
        self.assertFalse(lpe._cmds_are_equivalent(
            "app --a 1", "app --a 2"))


class TestCollectAllParamKeys(unittest.TestCase):
    def test_collects_unique_keys(self):
        params_list = [{"a": "1", "b": "2"}, {"b": "3", "c": "4"}]
        keys = lpe.collect_all_param_keys(params_list)
        self.assertEqual(keys, ["a", "b", "c"])

    def test_excludes_keys(self):
        params_list = [{"a": "1", "b": "2"}]
        keys = lpe.collect_all_param_keys(params_list, exclude=["b"])
        self.assertEqual(keys, ["a"])


# ============================================================
# 测试: 配置加载
# ============================================================

class TestLoadBenchConfig(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    def test_mode1_basic(self):
        path = _make_bench_yaml(self.tmp_dir)
        conf = lpe.load_bench_config(path)
        self.assertEqual(conf["mode"], 1)
        cases = conf["test_cases"]["cases"]
        self.assertEqual(len(cases), 1)
        self.assertEqual(cases[0], ((512, 256), (1, 5)))

    def test_mode1_combination_mode2(self):
        path = _make_bench_yaml(self.tmp_dir, {
            "test_cases": {
                "combination_mode": 2,
                "context": ["(512, 256)", "(1024, 1024)"],
                "batch_request": ["(1, 5)", "(2, 10)"],
            },
        })
        conf = lpe.load_bench_config(path)
        cases = conf["test_cases"]["cases"]
        self.assertEqual(len(cases), 2)
        self.assertEqual(cases[0], ((512, 256), (1, 5)))
        self.assertEqual(cases[1], ((1024, 1024), (2, 10)))

    def test_mode2_slo(self):
        path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "slo": {
                "criteria": {"ttft": "200ms", "tpot": "0.05s"},
                "context": [
                    {"context": "(4096, 1024)", "init_concurrent": 2},
                ],
            },
        })
        conf = lpe.load_bench_config(path)
        self.assertEqual(conf["mode"], 2)
        slo = conf["slo"]
        self.assertAlmostEqual(slo["criteria_parsed"]["ttft"], 0.2)
        self.assertAlmostEqual(slo["criteria_parsed"]["tpot"], 0.05)
        self.assertEqual(len(slo["contexts"]), 1)
        self.assertEqual(slo["contexts"][0]["context"], (4096, 1024))

    def test_mode2_engine_override_args_parsed(self):
        """engine_override_args 应被正确解析到 context 条目中"""
        path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "slo": {
                "criteria": {"ttft": "200ms", "tpot": "0.05s"},
                "context": [
                    {
                        "context": "(1024, 1024)",
                        "init_concurrent": 2,
                        "override_args": ["--common-arg 1"],
                        "engine_override_args": {
                            "vLLM": ["--vllm-arg 10"],
                            "SGLang": ["--sglang-arg 20", "!--remove-me"],
                        },
                    },
                    "(4096, 1024)",
                ],
            },
        })
        conf = lpe.load_bench_config(path)
        contexts = conf["slo"]["contexts"]
        # dict context with engine_override_args
        self.assertEqual(contexts[0]["override_args"], ["--common-arg 1"])
        self.assertEqual(contexts[0]["engine_override_args"]["vLLM"], ["--vllm-arg 10"])
        self.assertEqual(len(contexts[0]["engine_override_args"]["SGLang"]), 2)
        # simple string context: defaults to empty
        self.assertEqual(contexts[1]["override_args"], [])
        self.assertEqual(contexts[1]["engine_override_args"], {})

    def test_missing_required_field(self):
        conf = {"mode": 1, "tokenizer_path": "/fake", "url": "http://x"}
        path = os.path.join(self.tmp_dir, "bad.yaml")
        _write_yaml(path, conf)
        with self.assertRaises(ValueError):
            lpe.load_bench_config(path)

    def test_defaults_applied(self):
        path = _make_bench_yaml(self.tmp_dir)
        conf = lpe.load_bench_config(path)
        self.assertEqual(conf["result_dir"], "./results")
        self.assertFalse(conf["bench_only"])

    def test_result_file_prefix_from_model_name(self):
        path = _make_bench_yaml(self.tmp_dir, {"model_name": "org/my-model"})
        conf = lpe.load_bench_config(path)
        self.assertEqual(conf["result_file_prefix"], "org--my-model")

    def test_excel_prefix(self):
        path = _make_bench_yaml(self.tmp_dir, {
            "model_name": "m", "excel_prefix": "custom"
        })
        conf = lpe.load_bench_config(path)
        self.assertTrue(conf["result_file_prefix"].startswith("custom-"))


class TestLoadModelConfig(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    def test_basic_load(self):
        path = _make_model_yaml(self.tmp_dir)
        conf = lpe.load_model_config(path)
        engines = conf["engines"]
        self.assertEqual(len(engines), 1)
        self.assertEqual(engines[0]["engine_name"], "TestEngine")
        self.assertEqual(engines[0]["container_name"], "test-model-server")

    def test_exclude_params_merge(self):
        path = _make_model_yaml(self.tmp_dir, {
            "exclude_params": ["model"],
            "engines": [{
                "engine": "E1",
                "exclude_params": ["host"],
                "container_cmd": "docker run -itd --name c1 img:v1",
                "commands": [{"cmd": "serve --tp 4"}],
            }],
        })
        conf = lpe.load_model_config(path)
        excluded = conf["engines"][0]["exclude_params"]
        self.assertIn("model", excluded)
        self.assertIn("host", excluded)

    def test_ssh_inheritance(self):
        path = _make_model_yaml(self.tmp_dir, {
            "ssh_cmd": "ssh user@host",
            "engines": [{
                "engine": "E1",
                "container_cmd": "docker run -itd --name c1 img:v1",
                "commands": [{"cmd": "serve"}],
            }],
        })
        conf = lpe.load_model_config(path)
        self.assertEqual(conf["engines"][0]["ssh_cmd"], "ssh user@host")

    def test_missing_engines(self):
        path = os.path.join(self.tmp_dir, "bad.yaml")
        _write_yaml(path, {"engines": []})
        with self.assertRaises(ValueError):
            lpe.load_model_config(path)

    def test_image_name_parsed(self):
        path = _make_model_yaml(self.tmp_dir)
        conf = lpe.load_model_config(path)
        self.assertEqual(conf["engines"][0]["image_name"], "test-image:latest")


# ============================================================
# 测试: 指标提取与 SLO 判定
# ============================================================

class TestExtractAndParseMetrics(unittest.TestCase):
    def test_extract_from_list(self):
        raw = _fake_benchmark_result(ttft=0.1)
        metrics = lpe._extract_metrics(raw)
        self.assertIn("Average time to first token (s)", metrics)

    def test_parse_metrics(self):
        raw = _fake_benchmark_result(ttft=0.1, tpot=0.02, throughput=500)
        metrics = lpe._extract_metrics(raw)
        parsed = lpe._parse_metrics(metrics)
        self.assertEqual(parsed["ttft"], 0.1)
        self.assertEqual(parsed["tpot"], 0.02)
        self.assertEqual(parsed["throughput"], 500)
        self.assertEqual(parsed["comment"], "")

    def test_parse_metrics_with_failures(self):
        raw = _fake_benchmark_result(failed=3, total=20)
        metrics = lpe._extract_metrics(raw)
        parsed = lpe._parse_metrics(metrics)
        self.assertIn("失败请求数", parsed["comment"])

    def test_parse_metrics_none(self):
        self.assertIsNone(lpe._parse_metrics(None))
        self.assertIsNone(lpe._parse_metrics({}))


class TestCheckSloCriteria(unittest.TestCase):
    def test_pass(self):
        result = {"ttft": 0.1, "tpot": 0.03}
        criteria = {"ttft": 0.2, "tpot": 0.05}
        self.assertTrue(lpe.check_slo_criteria(result, criteria))

    def test_fail_ttft(self):
        result = {"ttft": 0.3, "tpot": 0.03}
        criteria = {"ttft": 0.2, "tpot": 0.05}
        self.assertFalse(lpe.check_slo_criteria(result, criteria))

    def test_none_result(self):
        self.assertFalse(lpe.check_slo_criteria(None, {"ttft": 0.2}))

    def test_missing_metric(self):
        result = {"ttft": 0.1}
        criteria = {"ttft": 0.2, "tpot": 0.05}
        self.assertFalse(lpe.check_slo_criteria(result, criteria))


# ============================================================
# 测试: 搜索算法
# ============================================================

class TestBinaryStep(unittest.TestCase):
    def test_pass_doubles(self):
        """达标且无上界 → 翻倍"""
        next_c, lb, ub, best, stop = lpe._binary_step(
            concurrent=2, is_valid=True,
            lower_bound=0, upper_bound=None,
            init_concurrent=2,
            best=None, batch_req=(2, 8),
            result={"ttft": 0.1, "tpot": 0.02})
        self.assertEqual(next_c, 4)
        self.assertEqual(lb, 2)
        self.assertIsNone(ub)
        self.assertIsNotNone(best)
        self.assertFalse(stop)

    def test_fail_binary_search(self):
        """未达标 → 二分"""
        next_c, lb, ub, best, stop = lpe._binary_step(
            concurrent=8, is_valid=False,
            lower_bound=4, upper_bound=None,
            init_concurrent=2,
            best=None, batch_req=(8, 32),
            result=None)
        self.assertEqual(next_c, 6)
        self.assertEqual(ub, 8)

    def test_converges(self):
        """区间收敛 → 停止"""
        next_c, lb, ub, best, stop = lpe._binary_step(
            concurrent=5, is_valid=True,
            lower_bound=4, upper_bound=6,
            init_concurrent=2,
            best=None, batch_req=(5, 20),
            result={"ttft": 0.1})
        self.assertTrue(stop)


class TestStopNStep(unittest.TestCase):
    def test_pass_increments(self):
        """达标且无上界 → +stop_n"""
        next_c, lb, ub, best, stop = lpe._stopn_step(
            concurrent=3, is_valid=True,
            lower_bound=0, upper_bound=None,
            stop_n=2,
            best=None, batch_req=(3, 12),
            result={"ttft": 0.1}, tested={3})
        self.assertEqual(next_c, 5)
        self.assertFalse(stop)

    def test_fail_decrements(self):
        """未达标且无下界 → -stop_n"""
        next_c, lb, ub, best, stop = lpe._stopn_step(
            concurrent=5, is_valid=False,
            lower_bound=0, upper_bound=None,
            stop_n=2,
            best=None, batch_req=(5, 20),
            result=None, tested={5})
        self.assertEqual(next_c, 3)
        self.assertFalse(stop)

    def test_already_tested_stops(self):
        """下一步已测试 → 停止"""
        next_c, lb, ub, best, stop = lpe._stopn_step(
            concurrent=3, is_valid=True,
            lower_bound=0, upper_bound=None,
            stop_n=2,
            best=None, batch_req=(3, 12),
            result={"ttft": 0.1}, tested={3, 5})
        self.assertTrue(stop)


# ============================================================
# 测试: ModelManager (mock subprocess / requests)
# ============================================================

class TestModelManager(unittest.TestCase):
    def _make_manager(self, **overrides):
        engine_conf = {
            "container_cmd": "docker run -itd --name mm-test img:v1",
            "stop_cmd": "docker stop mm-test && docker rm -f mm-test",
            "ssh_cmd": None,
            "container_name": "mm-test",
        }
        engine_conf.update(overrides)
        bench_conf = {
            "url": "http://localhost:8000/v1/chat/completions",
            "model_name": "test-model",
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 2},
        }
        return lpe.ModelManager(engine_conf, bench_conf)

    def test_build_full_cmd_normal(self):
        mgr = self._make_manager()
        full = mgr.build_full_cmd("python serve.py --port 8000")
        self.assertIn("docker run", full)
        self.assertIn("python serve.py --port 8000", full)

    def test_build_full_cmd_entrypoint_bash(self):
        mgr = self._make_manager(
            container_cmd="docker run --entrypoint bash img:v1")
        full = mgr.build_full_cmd("python serve.py")
        self.assertIn("-lc", full)

    @patch("subprocess.run")
    def test_cleanup_success(self, mock_run):
        mock_run.return_value = MagicMock(returncode=0, stdout="", stderr="")
        mgr = self._make_manager()
        with patch("time.sleep"):
            mgr.cleanup()
        mock_run.assert_called_once()

    @patch("subprocess.run")
    def test_start_model_success(self, mock_run):
        mock_run.return_value = MagicMock(returncode=0, stdout="ok", stderr="")
        mgr = self._make_manager()
        with patch("time.sleep"):
            success, logs = mgr.start_model("python serve.py")
        self.assertTrue(success)

    @patch("subprocess.run")
    def test_start_model_failure(self, mock_run):
        mock_run.return_value = MagicMock(returncode=1, stdout="", stderr="error")
        mgr = self._make_manager()
        with patch("time.sleep"):
            success, logs = mgr.start_model("python serve.py")
        self.assertFalse(success)

    @patch("requests.post")
    def test_health_check_pass(self, mock_post):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_post.return_value = mock_resp
        mgr = self._make_manager()
        with patch("time.sleep"):
            result = mgr.health_check()
        self.assertTrue(result)

    @patch("requests.post")
    def test_health_check_fail(self, mock_post):
        import requests as _req
        mock_post.side_effect = _req.ConnectionError("connection refused")
        mgr = self._make_manager()
        with patch("time.sleep"):
            result = mgr.health_check()
        self.assertFalse(result)

    @patch("subprocess.run")
    def test_ssh_cmd_wrapping(self, mock_run):
        mock_run.return_value = MagicMock(returncode=0, stdout="", stderr="")
        mgr = self._make_manager(ssh_cmd="ssh user@host")
        with patch("time.sleep"):
            mgr.cleanup()
        called_cmd = mock_run.call_args[0][0]
        self.assertTrue(called_cmd.startswith("ssh user@host"))


# ============================================================
# 测试: Excel 写入器
# ============================================================

class TestMode1ResultWriter(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()
        self.result_file = os.path.join(self.tmp_dir, "results", "test.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    def test_create_and_save(self):
        writer = lpe.Mode1ResultWriter(self.result_file)
        ws = writer.create_param_sheet(1, "docker run ... python serve.py")
        writer.append_result(ws, (512, 256), (1, 5), {
            "ttft": 0.1, "tpot": 0.02, "throughput": 500,
            "itl": 0.02, "duration": 10, "comment": "",
        })
        writer.append_result(ws, (512, 256), (2, 10), None)
        writer.finalize_sheet(ws)
        writer.save()
        self.assertTrue(os.path.exists(self.result_file))

    def test_multiple_sheets(self):
        writer = lpe.Mode1ResultWriter(self.result_file)
        ws1 = writer.create_param_sheet(1, "cmd1", "EngineA")
        ws2 = writer.create_param_sheet(2, "cmd2", "EngineA")
        self.assertEqual(ws1.title, "EngineA-参数1")
        self.assertEqual(ws2.title, "EngineA-参数2")
        writer.save()
        self.assertTrue(os.path.exists(self.result_file))


class TestMode2ResultWriter(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()
        self.result_file = os.path.join(self.tmp_dir, "results", "test-slo.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    def test_full_workflow(self):
        writer = lpe.Mode2ResultWriter(self.result_file, bench_only=False)
        writer.init_engine("TestEng", exclude_params=["model"])
        writer.register_command("TestEng", "serve --tp 4 --model m")
        ws = writer.create_detail_sheet("TestEng", 1, "docker run ... serve --tp 4")

        result = {"ttft": 0.1, "tpot": 0.02, "throughput": 500,
                  "itl": 0.02, "duration": 10, "comment": ""}
        writer.append_detail_result(ws, (4096, 1024), (2, 8), result)
        writer.add_summary_row(
            "TestEng", "test-model", "img:v1", 1, (4096, 1024),
            {"batch_req": (2, 8), "result": result})
        writer.finalize_detail_sheet(ws)
        writer.save()
        self.assertTrue(os.path.exists(self.result_file))

    def test_bench_only_workflow(self):
        writer = lpe.Mode2ResultWriter(self.result_file, bench_only=True)
        writer.init_engine("BenchOnly")
        writer.register_command("BenchOnly", "bench-only")
        ws = writer.create_detail_sheet("BenchOnly", 1, "bench-only")
        result = {"ttft": 0.1, "tpot": 0.02, "throughput": 500,
                  "itl": 0.02, "duration": 10, "comment": ""}
        writer.append_detail_result(ws, (4096, 1024), (3, 12), result)
        writer.add_summary_row(
            "BenchOnly", "test-model", "", 1, (4096, 1024),
            {"batch_req": (3, 12), "result": result})
        writer.save()
        self.assertTrue(os.path.exists(self.result_file))

    def test_failure_marker_and_logs(self):
        writer = lpe.Mode2ResultWriter(self.result_file)
        writer.init_engine("E1")
        ws = writer.create_detail_sheet("E1", 1, "cmd")
        writer.append_failure_marker(ws, (1024, 1024), "模型启动失败")
        writer.append_failure_logs(ws, "some error log\nline2", title="启动失败日志")
        writer.save()
        self.assertTrue(os.path.exists(self.result_file))

    def test_slo_no_pass_yellow_highlight(self):
        """SLO 全不达标时：明细表最小并发行黄色背景；汇总表黄色字体"""
        from openpyxl import load_workbook

        writer = lpe.Mode2ResultWriter(self.result_file, bench_only=True)
        writer.init_engine("EngX")
        writer.register_command("EngX", "bench-only")
        ws = writer.create_detail_sheet("EngX", 1, "bench-only")

        # 模拟两轮全不达标的结果（并发 2 和 4）
        r2 = {"ttft": 0.9, "tpot": 0.3, "throughput": 100, "itl": 0.1, "duration": 5, "comment": ""}
        r4 = {"ttft": 1.2, "tpot": 0.5, "throughput": 80,  "itl": 0.2, "duration": 5, "comment": ""}
        writer.append_detail_result(ws, (1024, 512), (2, 8), r2)
        writer.append_detail_result(ws, (1024, 512), (4, 16), r4)

        # min_tested = 并发2（最小）
        min_tested = {"batch_req": (2, 8), "result": r2}
        context_best_map = {(1024, 512): {"best": None, "min_tested": min_tested}}
        contexts = [{"context": [1024, 512]}]
        writer.highlight_best_rows(ws, context_best_map, contexts)

        # 验证并发=2 的行（row 2）填充黄色
        yellow = "FFFFFF00"
        row2_fill = ws.cell(row=2, column=3).fill.fgColor.rgb  # column 3 = 并发数
        self.assertEqual(row2_fill, yellow)

        # 并发=4 的行不应是黄色
        row3_fill = ws.cell(row=3, column=3).fill.fgColor.rgb
        self.assertNotEqual(row3_fill, yellow)

        # 汇总表：slo_no_pass 行应为黄色字体
        writer.add_summary_row("EngX", "m", "", 1, (1024, 512), min_tested, slo_no_pass=True)
        writer.write_all_summaries()
        writer.save()

        wb = load_workbook(self.result_file)
        summary_ws = wb["汇总-EngX"]
        # row 2 (first data row) should have yellow font
        font_color = summary_ws.cell(row=2, column=1).font.color.rgb
        self.assertEqual(font_color, "FFFFCC00")


# ============================================================
# 测试: 端到端 run_mode1 / run_mode2 (mock evalscope)
# ============================================================

class TestRunMode1E2E(unittest.TestCase):
    """端到端测试模式1，mock 掉 evalscope 和模型管理"""

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    @patch("llm_perf_eval.run_perf_benchmark")
    def test_bench_only(self, mock_bench):
        """bench_only 模式应跳过模型启停，直接运行压测"""
        mock_bench.return_value = _fake_benchmark_result()

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "bench_only": True,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "test_cases": {
                "combination_mode": 1,
                "context": ["(512, 256)"],
                "batch_request": ["(1, 5)", "(2, 10)"],
            },
        })
        bench_conf = lpe.load_bench_config(bench_path)
        model_conf = {
            "engines": [{
                "engine_name": "bench-only",
                "container_cmd": "",
                "stop_cmd": "",
                "ssh_cmd": None,
                "container_name": None,
                "image_name": "",
                "exclude_params": [],
                "commands": [{"cmd": "bench-only"}],
            }],
        }

        lpe.run_mode1(bench_conf, model_conf)
        self.assertEqual(mock_bench.call_count, 2)

        # 验证 Excel 文件生成
        result_dir = os.path.join(self.tmp_dir, "results")
        xlsx_files = [f for f in os.listdir(result_dir) if f.endswith(".xlsx")]
        self.assertEqual(len(xlsx_files), 1)


class TestBenchOnlyHealthcheck(unittest.TestCase):
    """bench_only 模式下可选健康检查"""

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    def _bench_only_model_conf(self):
        return {
            "engines": [{
                "engine_name": "bench-only",
                "container_cmd": "",
                "stop_cmd": "",
                "ssh_cmd": None,
                "container_name": None,
                "image_name": "",
                "exclude_params": [],
                "commands": [{"cmd": "bench-only"}],
            }],
        }

    @patch("llm_perf_eval.run_perf_benchmark")
    @patch("requests.post")
    def test_mode1_bench_only_healthcheck_pass(self, mock_post, mock_bench):
        """bench_only + healthcheck=true，健康检查通过后正常压测"""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_post.return_value = mock_resp
        mock_bench.return_value = _fake_benchmark_result()

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "bench_only": True,
            "bench_only_healthcheck": True,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
        })
        bench_conf = lpe.load_bench_config(bench_path)
        with patch("time.sleep"):
            lpe.run_mode1(bench_conf, self._bench_only_model_conf())
        self.assertTrue(mock_bench.called)

    @patch("llm_perf_eval.run_perf_benchmark")
    @patch("requests.post")
    def test_mode1_bench_only_healthcheck_fail_skips(self, mock_post, mock_bench):
        """bench_only + healthcheck=true，健康检查失败则跳过压测"""
        import requests as _req
        mock_post.side_effect = _req.ConnectionError("refused")

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "bench_only": True,
            "bench_only_healthcheck": True,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
        })
        bench_conf = lpe.load_bench_config(bench_path)
        with patch("time.sleep"):
            lpe.run_mode1(bench_conf, self._bench_only_model_conf())
        # 健康检查失败，不应调用 benchmark
        self.assertFalse(mock_bench.called)

    @patch("llm_perf_eval.run_perf_benchmark")
    @patch("requests.post")
    def test_mode2_bench_only_healthcheck_fail_skips(self, mock_post, mock_bench):
        """mode2 bench_only + healthcheck=true，健康检查失败则跳过"""
        import requests as _req
        mock_post.side_effect = _req.ConnectionError("refused")

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "bench_only": True,
            "bench_only_healthcheck": True,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
            "slo": {
                "criteria": {"ttft": "0.2s", "tpot": "0.05s"},
                "search_method": "binary",
                "init_concurrent": 2,
                "request_multiplier": 4,
                "context": [{"context": "(1024, 512)", "init_concurrent": 2}],
            },
        })
        bench_conf = lpe.load_bench_config(bench_path)
        with patch("time.sleep"):
            lpe.run_mode2(bench_conf, self._bench_only_model_conf())
        self.assertFalse(mock_bench.called)


class TestRunMode2E2E(unittest.TestCase):
    """端到端测试模式2 (SLO)，mock evalscope"""

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    @patch("llm_perf_eval.run_perf_benchmark")
    def test_bench_only_slo_binary(self, mock_bench):
        """bench_only + binary 搜索：验证搜索收敛并产出 Excel"""
        call_count = [0]

        def _side_effect(args):
            call_count[0] += 1
            # 前 3 次达标，第 4 次开始失败
            if call_count[0] <= 3:
                return _fake_benchmark_result(ttft=0.05, tpot=0.02)
            return _fake_benchmark_result(ttft=0.5, tpot=0.1)

        mock_bench.side_effect = _side_effect

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "bench_only": True,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "slo": {
                "criteria": {"ttft": "0.2s", "tpot": "0.05s"},
                "search_method": "binary",
                "init_concurrent": 2,
                "request_multiplier": 4,
                "context": [
                    {"context": "(1024, 512)", "init_concurrent": 2},
                ],
            },
        })
        bench_conf = lpe.load_bench_config(bench_path)
        model_conf = {
            "engines": [{
                "engine_name": "bench-only",
                "container_cmd": "",
                "stop_cmd": "",
                "ssh_cmd": None,
                "container_name": None,
                "image_name": "",
                "exclude_params": [],
                "commands": [{"cmd": "bench-only"}],
            }],
        }

        lpe.run_mode2(bench_conf, model_conf)
        self.assertGreater(call_count[0], 1)

        result_dir = os.path.join(self.tmp_dir, "results")
        xlsx_files = [f for f in os.listdir(result_dir) if f.endswith(".xlsx")]
        self.assertEqual(len(xlsx_files), 1)

    @patch("llm_perf_eval.run_perf_benchmark")
    def test_bench_only_slo_stopn(self, mock_bench):
        """bench_only + stopN 搜索"""
        call_count = [0]

        def _side_effect(args):
            call_count[0] += 1
            if call_count[0] <= 2:
                return _fake_benchmark_result(ttft=0.05, tpot=0.02)
            return _fake_benchmark_result(ttft=0.5, tpot=0.1)

        mock_bench.side_effect = _side_effect

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "bench_only": True,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "slo": {
                "criteria": {"ttft": "0.2s", "tpot": "0.05s"},
                "search_method": "stopN",
                "stop_n": 2,
                "init_concurrent": 2,
                "request_multiplier": 4,
                "context": ["(1024, 512)"],
            },
        })
        bench_conf = lpe.load_bench_config(bench_path)
        model_conf = {
            "engines": [{
                "engine_name": "bench-only",
                "container_cmd": "",
                "stop_cmd": "",
                "ssh_cmd": None,
                "container_name": None,
                "image_name": "",
                "exclude_params": [],
                "commands": [{"cmd": "bench-only"}],
            }],
        }

        lpe.run_mode2(bench_conf, model_conf)
        self.assertGreater(call_count[0], 1)

    @patch("llm_perf_eval.run_perf_benchmark")
    @patch("subprocess.run")
    @patch("requests.post")
    def test_full_mode2_with_model_lifecycle(self, mock_post, mock_run, mock_bench):
        """完整模式2：含模型启停、健康检查（全 mock）"""
        mock_run.return_value = MagicMock(returncode=0, stdout="ok", stderr="")
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_post.return_value = mock_resp

        call_count = [0]

        def _side_effect(args):
            call_count[0] += 1
            if call_count[0] <= 2:
                return _fake_benchmark_result(ttft=0.05, tpot=0.02)
            return _fake_benchmark_result(ttft=0.5, tpot=0.1)

        mock_bench.side_effect = _side_effect

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "bench_only": False,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
            "slo": {
                "criteria": {"ttft": "0.2s", "tpot": "0.05s"},
                "search_method": "binary",
                "init_concurrent": 2,
                "request_multiplier": 4,
                "context": [
                    {"context": "(1024, 512)", "init_concurrent": 2},
                ],
            },
        })
        model_path = _make_model_yaml(self.tmp_dir)

        bench_conf = lpe.load_bench_config(bench_path)
        model_conf = lpe.load_model_config(model_path)

        with patch("time.sleep"):
            lpe.run_mode2(bench_conf, model_conf)

        # 验证 subprocess.run 被调用（模型启停）
        self.assertTrue(mock_run.called)
        self.assertGreater(call_count[0], 0)

    @patch("llm_perf_eval.run_perf_benchmark")
    @patch("subprocess.run")
    @patch("requests.post")
    def test_mode2_engine_override_args_e2e(self, mock_post, mock_run, mock_bench):
        """E2E: engine_override_args 使不同引擎收到不同的启动命令"""
        mock_run.return_value = MagicMock(returncode=0, stdout="ok", stderr="")
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_post.return_value = mock_resp
        mock_bench.return_value = _fake_benchmark_result(ttft=0.5, tpot=0.1)

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "mode": 2,
            "bench_only": False,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
            "slo": {
                "criteria": {"ttft": "0.2s", "tpot": "0.05s"},
                "search_method": "binary",
                "init_concurrent": 2,
                "request_multiplier": 4,
                "context": [{
                    "context": "(1024, 512)",
                    "init_concurrent": 2,
                    "override_args": ["--common 1"],
                    "engine_override_args": {
                        "EngA": ["--enga-only 10"],
                        "EngB": ["--engb-only 20"],
                    },
                }],
            },
        })
        model_conf = {
            "engines": [
                {
                    "engine_name": "EngA",
                    "container_cmd": "docker run -itd --name ea img:v1",
                    "stop_cmd": "docker stop ea",
                    "ssh_cmd": None,
                    "container_name": "ea",
                    "image_name": "img:v1",
                    "exclude_params": [],
                    "commands": [{"cmd": "serve --tp 4 --host 0.0.0.0 --port 8000"}],
                },
                {
                    "engine_name": "EngB",
                    "container_cmd": "docker run -itd --name eb img:v2",
                    "stop_cmd": "docker stop eb",
                    "ssh_cmd": None,
                    "container_name": "eb",
                    "image_name": "img:v2",
                    "exclude_params": [],
                    "commands": [{"cmd": "serve --tp 8 --host 0.0.0.0 --port 8000"}],
                },
            ],
        }

        bench_conf = lpe.load_bench_config(bench_path)

        # 记录 start_model 收到的实际命令
        started_cmds = []
        orig_start = lpe.ModelManager.start_model
        def _capture_start(self, cmd):
            started_cmds.append(cmd)
            return True, "ok"
        with patch("time.sleep"), \
             patch.object(lpe.ModelManager, "start_model", _capture_start), \
             patch.object(lpe.ModelManager, "stop_model"), \
             patch.object(lpe.ModelManager, "health_check", return_value=True), \
             patch.object(lpe.ModelManager, "cleanup"):
            lpe.run_mode2(bench_conf, model_conf)

        # EngA 应收到 --common 1 --enga-only 10
        self.assertTrue(len(started_cmds) >= 2)
        enga_cmd = lpe.parse_model_cmd_args(started_cmds[0])
        engb_cmd = lpe.parse_model_cmd_args(started_cmds[1])
        self.assertEqual(enga_cmd.get("common"), "1")
        self.assertEqual(enga_cmd.get("enga-only"), "10")
        self.assertNotIn("engb-only", enga_cmd)
        # EngB 应收到 --common 1 --engb-only 20
        self.assertEqual(engb_cmd.get("common"), "1")
        self.assertEqual(engb_cmd.get("engb-only"), "20")
        self.assertNotIn("enga-only", engb_cmd)


class TestRunMode1WithModelLifecycle(unittest.TestCase):
    """模式1 含模型启停"""

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp_dir)

    @patch("llm_perf_eval.run_perf_benchmark")
    @patch("subprocess.run")
    @patch("requests.post")
    def test_mode1_full(self, mock_post, mock_run, mock_bench):
        mock_run.return_value = MagicMock(returncode=0, stdout="ok", stderr="")
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_post.return_value = mock_resp
        mock_bench.return_value = _fake_benchmark_result()

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "bench_only": False,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
        })
        model_path = _make_model_yaml(self.tmp_dir)

        bench_conf = lpe.load_bench_config(bench_path)
        model_conf = lpe.load_model_config(model_path)

        with patch("time.sleep"):
            lpe.run_mode1(bench_conf, model_conf)

        self.assertTrue(mock_bench.called)
        result_dir = os.path.join(self.tmp_dir, "results")
        xlsx_files = [f for f in os.listdir(result_dir) if f.endswith(".xlsx")]
        self.assertEqual(len(xlsx_files), 1)

    @patch("subprocess.run")
    def test_mode1_model_start_failure(self, mock_run):
        """模型启动失败时应标记失败并继续"""
        mock_run.return_value = MagicMock(returncode=1, stdout="", stderr="launch error")

        bench_path = _make_bench_yaml(self.tmp_dir, {
            "bench_only": False,
            "warmup": {"enable": False},
            "result_dir": os.path.join(self.tmp_dir, "results"),
            "healthcheck": {"initial_delay": 0, "interval": 0, "retry_count": 1},
        })
        model_path = _make_model_yaml(self.tmp_dir)

        bench_conf = lpe.load_bench_config(bench_path)
        model_conf = lpe.load_model_config(model_path)

        with patch("time.sleep"):
            lpe.run_mode1(bench_conf, model_conf)

        result_dir = os.path.join(self.tmp_dir, "results")
        xlsx_files = [f for f in os.listdir(result_dir) if f.endswith(".xlsx")]
        self.assertEqual(len(xlsx_files), 1)


# ============================================================
# 测试: Warmup
# ============================================================

class TestWarmup(unittest.TestCase):
    @patch("llm_perf_eval.run_perf_benchmark")
    def test_warmup_success(self, mock_bench):
        mock_bench.return_value = _fake_benchmark_result()
        bench_conf = {
            "model_name": "m", "url": "http://x", "tokenizer_path": "/t",
            "warmup": {"enable": True, "context": (128, 128), "parallel": 5, "number": 10},
        }
        result = lpe.run_warmup(bench_conf)
        self.assertTrue(result)
        mock_bench.assert_called_once()

    @patch("llm_perf_eval.run_perf_benchmark")
    def test_warmup_failure_continues(self, mock_bench):
        mock_bench.side_effect = Exception("warmup failed")
        bench_conf = {
            "model_name": "m", "url": "http://x", "tokenizer_path": "/t",
            "warmup": {},
        }
        result = lpe.run_warmup(bench_conf)
        self.assertFalse(result)


# ============================================================
# 测试: _apply_defaults
# ============================================================

class TestApplyDefaults(unittest.TestCase):
    def test_fills_missing(self):
        conf = {"a": 1}
        lpe._apply_defaults(conf, {"a": 99, "b": 2})
        self.assertEqual(conf["a"], 1)
        self.assertEqual(conf["b"], 2)

    def test_fills_none(self):
        conf = {"a": None}
        lpe._apply_defaults(conf, {"a": 5})
        self.assertEqual(conf["a"], 5)


# ============================================================
# 测试: container / image 解析
# ============================================================

class TestContainerParsing(unittest.TestCase):
    def test_parse_container_name(self):
        cmd = "docker run -itd --name my-container --gpus all img:v1"
        self.assertEqual(lpe._parse_container_name(cmd), "my-container")

    def test_parse_container_name_missing(self):
        self.assertIsNone(lpe._parse_container_name("docker run img:v1"))

    def test_parse_image_name(self):
        cmd = "docker run -itd --name c1 --gpus all registry.example.com/img:v1"
        self.assertEqual(lpe._parse_image_name(cmd), "registry.example.com/img:v1")

    def test_parse_image_name_multiline(self):
        cmd = "docker run -itd \\\n--name c1 \\\nimg:v2"
        self.assertEqual(lpe._parse_image_name(cmd), "img:v2")


# ============================================================
# 运行
# ============================================================

if __name__ == "__main__":
    unittest.main()
